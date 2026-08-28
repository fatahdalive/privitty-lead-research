import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

# ============================================================
# FILE PATHS
# ============================================================

INPUT_FILE = "data/shortlisted_companies.csv"
OUTPUT_FILE = "data/final_leads.xlsx"


# ============================================================
# LOAD SHORTLIST
# ============================================================

print("=" * 60)
print("PRIVITTY LEAD ENRICHMENT")
print("=" * 60)

print("\nLoading shortlisted companies...")

df = pd.read_csv(INPUT_FILE)

print(f"Companies loaded: {len(df)}")


# ============================================================
# CLEAN EMPTY VALUES
# ============================================================

df = df.fillna("")


# ============================================================
# HELPER FUNCTION
# ============================================================

def clean_value(value):
    """
    Converts empty/NaN-like values into a clean blank.
    """
    value = str(value).strip()

    if value.lower() in ["nan", "none", "null"]:
        return ""

    return value


# ============================================================
# CLEAN IMPORTANT FIELDS
# ============================================================

for column in [
    "Name",
    "Email",
    "Phone",
    "Website",
    "Address",
    "Categories",
    "Relevant Countries",
    "Quality"
]:

    if column in df.columns:
        df[column] = df[column].apply(clean_value)


# ============================================================
# GENERATE LEAD TYPE
# ============================================================

def determine_lead_type(row):

    categories = str(row.get("Categories", "")).lower()

    if any(word in categories for word in [
        "finance",
        "bank",
        "banking",
        "insurance",
        "credit"
    ]):
        return "Financial / Regulated"

    if any(word in categories for word in [
        "health",
        "healthcare",
        "medical",
        "hospital"
    ]):
        return "Healthcare / Sensitive Data"

    if any(word in categories for word in [
        "telecom",
        "telecommunication"
    ]):
        return "Telecommunications"

    if any(word in categories for word in [
        "cloud",
        "saas",
        "security",
        "identity"
    ]):
        return "Technology / Data"

    if any(word in categories for word in [
        "advertising",
        "ads",
        "marketing",
        "social media"
    ]):
        return "Advertising / Consumer Data"

    if any(word in categories for word in [
        "ecommerce",
        "retail",
        "marketplace"
    ]):
        return "Commerce / Consumer Data"

    return "Other"


df["Lead Type"] = df.apply(determine_lead_type, axis=1)


# ============================================================
# GENERATE OUTREACH READINESS
# ============================================================

def outreach_readiness(row):

    available = 0

    if clean_value(row.get("Email", "")):
        available += 1

    if clean_value(row.get("Phone", "")):
        available += 1

    if clean_value(row.get("Website", "")):
        available += 1

    if available == 3:
        return "High"

    elif available == 2:
        return "Medium"

    elif available == 1:
        return "Low"

    return "Unavailable"


df["Outreach Readiness"] = df.apply(
    outreach_readiness,
    axis=1
)


# ============================================================
# GENERATE WHY SELECTED
# ============================================================

def generate_reason(row):

    reasons = []

    categories = str(row.get("Categories", "")).lower()

    # Category reason
    if any(word in categories for word in [
        "finance",
        "bank",
        "banking",
        "insurance",
        "credit"
    ]):
        reasons.append("handles financial or regulated data")

    elif any(word in categories for word in [
        "health",
        "healthcare",
        "medical",
        "hospital"
    ]):
        reasons.append("potentially handles sensitive healthcare data")

    elif any(word in categories for word in [
        "telecom",
        "telecommunication"
    ]):
        reasons.append("handles customer telecommunications data")

    elif any(word in categories for word in [
        "cloud",
        "saas",
        "security",
        "identity"
    ]):
        reasons.append("operates in a data-intensive technology sector")

    elif any(word in categories for word in [
        "advertising",
        "ads",
        "marketing",
        "social media"
    ]):
        reasons.append("likely processes consumer or advertising data")

    elif any(word in categories for word in [
        "ecommerce",
        "retail",
        "marketplace"
    ]):
        reasons.append("handles consumer or transaction data")

    # Quality
    if str(row.get("Quality", "")).lower() == "verified":
        reasons.append("verified company information")

    # Contact information
    if clean_value(row.get("Email", "")):
        reasons.append("public email available")

    if clean_value(row.get("Phone", "")):
        reasons.append("public phone available")

    if clean_value(row.get("Website", "")):
        reasons.append("website available")

    if not reasons:
        return "Selected based on lead score"

    return "; ".join(reasons)


df["Why Selected"] = df.apply(
    generate_reason,
    axis=1
)


# ============================================================
# REORDER COLUMNS
# ============================================================

preferred_columns = [
    "Name",
    "Lead Score",
    "Priority",
    "Lead Type",
    "Outreach Readiness",
    "Reason",
    "Why Selected",
    "Email",
    "Phone",
    "Website",
    "Categories",
    "Relevant Countries",
    "Quality",
    "Address",
    "Sources",
    "Directory URL",
    "JSON URL",
    "Slug"
]

# Keep only columns that actually exist
final_columns = [
    column for column in preferred_columns
    if column in df.columns
]

df = df[final_columns]


# ============================================================
# SORT BY LEAD SCORE
# ============================================================

if "Lead Score" in df.columns:

    df["Lead Score"] = pd.to_numeric(
        df["Lead Score"],
        errors="coerce"
    ).fillna(0)

    df = df.sort_values(
        by="Lead Score",
        ascending=False
    )


# ============================================================
# EXPORT TO EXCEL
# ============================================================

print("\nCreating Excel file...")

df.to_excel(
    OUTPUT_FILE,
    index=False,
    sheet_name="Privitty Leads"
)


# ============================================================
# FORMAT EXCEL
# ============================================================

print("Formatting Excel workbook...")

workbook = load_workbook(OUTPUT_FILE)

worksheet = workbook["Privitty Leads"]


# Freeze header row
worksheet.freeze_panes = "A2"


# Enable filters / Excel table
last_row = worksheet.max_row
last_column = worksheet.max_column

last_column_letter = get_column_letter(last_column)

table_range = f"A1:{last_column_letter}{last_row}"

table = Table(
    displayName="PrivittyLeadTable",
    ref=table_range
)

table_style = TableStyleInfo(
    name="TableStyleMedium2",
    showFirstColumn=False,
    showLastColumn=False,
    showRowStripes=True,
    showColumnStripes=False
)

table.tableStyleInfo = table_style

worksheet.add_table(table)


# Format header
for cell in worksheet[1]:

    cell.font = Font(
        bold=True
    )

    cell.alignment = Alignment(
        horizontal="center",
        vertical="center"
    )


# Wrap text
for row in worksheet.iter_rows():

    for cell in row:

        cell.alignment = Alignment(
            vertical="top",
            wrap_text=True
        )


# Set useful column widths
column_widths = {
    "A": 28,   # Name
    "B": 12,   # Score
    "C": 12,   # Priority
    "D": 28,   # Lead Type
    "E": 18,   # Outreach
    "F": 35,   # Reason
    "G": 60,   # Why Selected
    "H": 32,   # Email
    "I": 20,   # Phone
    "J": 35,   # Website
    "K": 35,   # Categories
    "L": 25,   # Countries
    "M": 15,   # Quality
    "N": 40,   # Address
    "O": 40,   # Sources
    "P": 45,   # Directory URL
    "Q": 45,   # JSON URL
    "R": 25    # Slug
}

for column, width in column_widths.items():

    worksheet.column_dimensions[column].width = width


# Header row height
worksheet.row_dimensions[1].height = 30


# Save final workbook
workbook.save(OUTPUT_FILE)


# ============================================================
# SUMMARY
# ============================================================

print("\n" + "=" * 60)
print("ENRICHMENT COMPLETED SUCCESSFULLY")
print("=" * 60)

print(f"Companies processed : {len(df)}")
print(f"Excel file          : {OUTPUT_FILE}")

if "Priority" in df.columns:

    print("\nPriority breakdown:")

    print(
        df["Priority"].value_counts()
    )

if "Lead Type" in df.columns:

    print("\nLead type breakdown:")

    print(
        df["Lead Type"].value_counts().head(10)
    )

print("\nFinal file is ready!")